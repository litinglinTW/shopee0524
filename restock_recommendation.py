
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import datetime
import traceback

# 清理商品名稱
def ultimate_clean_name(name):
    if not isinstance(name, str):
        return ""
    name = re.sub(r"𝐁𝐥𝐮𝐞\.𝐖𝐨𝐧𝐝𝐞𝐫", "", name)
    name = re.sub(r"[｜|]", "", name)
    name = name.replace("現貨", "").strip()
    name = re.sub(r"[A-Da-d]\d{3}.*$", "", name).strip()
    return name

# 擷取商品編號
def extract_code(text):
    if not isinstance(text, str):
        return ""
    match = re.search(r"[A-Da-d]\d{3}", text)
    return match.group(0).upper() if match else ""

# 主功能：進貨建議產生器
def generate_restock_recommendation(inventory_file, sales_file):
    result = {"success": False, "msg": "", "data": None, "bytes": None}

    try:
        df_inventory = pd.read_excel(inventory_file)
        df_sales = pd.read_excel(sales_file)

        # 欄位準備
        df_inventory["商品ID"] = df_inventory["et_title_product_id"].astype(str)
        df_inventory["商品規格ID"] = df_inventory["et_title_variation_id"].astype(str)
        df_inventory["商品名稱原始"] = df_inventory["et_title_product_name"]
        df_inventory["商品規格來源"] = df_inventory["et_title_variation_name"]
        df_inventory["庫存數量"] = pd.to_numeric(df_inventory["et_title_variation_stock"], errors="coerce").fillna(0).astype(int)

        df_sales["商品ID"] = df_sales["商品ID"].astype(str)
        df_sales["商品規格ID"] = df_sales["商品規格ID"].astype(str)
        df_sales["銷售數量"] = df_sales["數量(全部訂單)"]

        # 🔧 避免商品規格重複（若銷售報表意外含此欄）
        if "商品規格" in df_sales.columns:
            df_sales.drop(columns=["商品規格"], inplace=True)


        # 過濾銷售數量 > 3 且 > 庫存數量
        filtered_sales = df_sales[df_sales["銷售數量"] > 3]
        merged = pd.merge(df_inventory, filtered_sales, on=["商品ID", "商品規格ID"], how="inner")
        merged = merged[merged["銷售數量"] > merged["庫存數量"]]

        # 若合併後已有商品規格欄位，避免重複命名造成錯誤
        if "商品規格" in merged.columns:
            merged.drop(columns=["商品規格"], inplace=True)

        # ✅ 修正：商品規格重新命名為簡潔欄名
        merged = merged.rename(columns={"商品規格來源": "商品規格"})


        # 清理與提取
        merged["商品名稱"] = merged["商品名稱原始"].apply(ultimate_clean_name)
        merged["商品編號"] = merged["商品名稱原始"].apply(extract_code)

        final_df = merged.loc[:, [
            "商品編號", "商品名稱", "商品規格", "銷售數量", "庫存數量"
        ]].copy()

        # 最終確認：刪除所有欄位名稱重複的情況（安全檢查）
        final_df = final_df.loc[:, ~final_df.columns.duplicated()]
        final_df["建議補貨量"] = final_df["銷售數量"] * 2 - final_df["庫存數量"]


        # 轉成 Excel
        output = BytesIO()
        wb = load_workbook("assets/restock_recommendation_template.xlsx")
        ws = wb.active

        # 清除舊資料（從第 2 列開始）
        ws.delete_rows(2, ws.max_row - 1)

        # 寫入資料從第 2 列開始
        for row in dataframe_to_rows(final_df, index=False, header=False):
            if any(row):  # 避免寫入全空白列
                current_row = ws.max_row + 1
                ws.append([
                    row[0],  # 商品編號
                    row[1],  # 商品名稱
                    row[2],  # 商品規格
                    row[3],  # 銷售數量
                    row[4],  # 庫存數量
                    f"=D{current_row}+D{current_row}-E{current_row}"  # 建議補貨量公式
                ])




        today = datetime.today().strftime("%Y-%m-%d")
        ws["A1"] = f"進貨建議清單 – {today}"
        ws["A1"].font = Font(size=14, bold=True)

        wb.save(output)
        output.seek(0)

        result.update({
            "success": True,
            "data": final_df,
            "bytes": output,
            "msg": "✅ 成功產生建議報表"
        })
        return result

    except Exception as e:
        error_detail = traceback.format_exc()
        
        debug_columns = ""
        try:
            debug_columns = f"\n\n👉 merged.columns:\n{list(merged.columns)}"
        except:
            debug_columns = "\n\n⚠️ 無法印出 merged.columns，可能尚未定義"

        result["msg"] = f"🚨 發生錯誤：{e}\n\n詳情：\n{error_detail}{debug_columns}"
        return result

