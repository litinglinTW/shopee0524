import re
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from io import BytesIO

# === 基本設定 ===
'''
FOLDER = os.path.expanduser("~/Documents/Blue.Wonder財報數據庫")
TEMPLATE_FILENAME = "BlueWonder_財報模板.xlsx"
PDF_PREFIX = "monthly_report_"
EXCEL_PREFIX = "[export_report]sales_overview_"
'''

DASHBOARD_SHEET = "Dashboard"

def extract_summary_from_pdf(pdf_bytes):
    reader = PdfReader(pdf_bytes)
    text = "\n".join(page.extract_text() for page in reader.pages)

    def extract(patterns, text, label):
        if isinstance(patterns, str):
            patterns = [patterns]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return int(match.group(1).replace(",", ""))
        print(f"⚠️ 找不到欄位：{label}，將預設為 0，請確認是否需手動填寫。")
        return 0

    summary = {
        "商品價格": extract([r"商品價格\s*([\d,]+)", r"買家支付商品金額\s*([\d,]+)"], text, "商品價格"),
        "退款金額": extract(r"退款金額\s*-?([\d,]+)", text, "退款金額"),
        "成交手續費": extract(r"成交手續費\s*-?([\d,]+)", text, "成交手續費"),
        "其他服務費": extract(r"其他服務費\s*-?([\d,]+)", text, "其他服務費"),
        "金流與系統處理費": extract(r"金流與系統處理費\s*-?([\d,]+)", text, "金流與系統處理費"),
        "優惠券與補貼": extract(r"優惠券與補貼\s*-?([\d,]+)", text, "優惠券與補貼"),
        "賣家運費總支付": extract(r"賣家運費總支付\s*-?([\d,]+)", text, "賣家運費總支付"),
        "訂單調整退款": extract(r"退貨退款\s*-?([\d,]+)", text, "訂單調整退款"),
    }
    if summary["成交手續費"] is not None and summary["其他服務費"] is not None:
        summary["平台手續費"] = summary["成交手續費"] + summary["其他服務費"]
    else:
        summary["平台手續費"] = None
    return summary, text   # 回傳 text 方便 debug
    


def extract_daily_sales_data(excel_bytes):
    wb = load_workbook(excel_bytes, data_only=True)
    ws = wb.active

    daily_sales = []
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=6):
        date_cell, _, _, _, _, amount_cell = row
        if date_cell.value and amount_cell.value:
            daily_sales.append((date_cell.value, amount_cell.value))

    return daily_sales


# === 建立月份名稱 ===
def get_month_from_filename(filename):
    match = re.search(r"(\d{6})\d{2}", filename)
    if not match:
        raise ValueError("無法從檔名擷取月份，請確認檔名格式為 monthly_report_YYYYMMDD.pdf")
    yyyymm = match.group(1)
    return f"{yyyymm[:4]}-{yyyymm[4:]}"



# === 主程式 ===
def run_financial_report(pdf_file, excel_file, template_file):
    result = {"success": False, "msg": "", "sheet": None}
    try:
        month_name = get_month_from_filename(pdf_file.name)
        summary, pdf_text = extract_summary_from_pdf(pdf_file)
        sales_df = extract_daily_sales_data(excel_file)

        wb = load_workbook(template_file)

        if month_name not in wb.sheetnames:
            result["msg"] = f"{month_name} 分頁不存在，請先在 Excel 裡手動建立"
            return result

        if DASHBOARD_SHEET not in wb.sheetnames:
            result["msg"] = f"找不到 Dashboard 分頁，請確認模板中有名為 '{DASHBOARD_SHEET}' 的工作表"
            return result

        new_sheet = wb[month_name]

        new_sheet["B2"] = summary.get("商品價格", 0)
        new_sheet["B4"] = summary.get("退款金額", 0)
        new_sheet["B5"] = summary.get("平台手續費", 0)
        new_sheet["B6"] = summary.get("金流與系統處理費", 0)
        new_sheet["B10"] = summary.get("優惠券與補貼", 0)
        new_sheet["B25"] = summary.get("賣家運費總支付", 0)
        new_sheet["B30"] = summary.get("訂單調整退款", 0)
        new_sheet["B3"] = load_workbook(excel_file, data_only=True).active["E2"].value

        for i, (day, amount) in enumerate(sales_df, start=3):
            new_sheet[f"Q{i}"] = day
            new_sheet[f"R{i}"] = float(str(amount).replace(",", ""))

        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        result["success"] = True
        result["msg"] = f"✅ 成功產出 {month_name} 分頁的財報"
        result["sheet"] = month_name
        result["report_bytes"] = output_buffer
        
        # 在 try 區塊結尾前（success 為 True 的區塊）
        result["summary"] = summary[0]  # summary 是 tuple: (summary_dict, pdf_text)
        result["sales"] = sales_df
        result["pdf_raw"] = pdf_text  # 加上 PDF 原始文字內容 debug 用
        
          

        return result
    except Exception as e:
        result["msg"] = f"🚨 財報分析失敗：{e}"
        return result

'''

＃本機測試檔案路徑固定版本：


# === 尋找最新 PDF 檔案 ===
def find_latest_pdf(folder):
    pdfs = [f for f in os.listdir(folder) if f.startswith(PDF_PREFIX) and f.endswith(".pdf")]
    if not pdfs:
        raise FileNotFoundError("找不到 monthly_report_ 開頭的 PDF")
    pdfs.sort(reverse=True)
    return os.path.join(folder, pdfs[0])

# === 尋找最新 Excel 銷售報表 ===
def find_latest_excel(folder):
    excels = [f for f in os.listdir(folder) if f.startswith(EXCEL_PREFIX) and f.endswith(".xlsx")]
    if not excels:
        raise FileNotFoundError("找不到 sales_overview_ 開頭的 Excel")
    excels.sort(reverse=True)
    return os.path.join(folder, excels[0])

'''


