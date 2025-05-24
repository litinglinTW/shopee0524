
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import difflib
import pandas as pd
from datetime import datetime

def generate_purchase_report(purchase_file, template_file, mapping_file):
    wb = load_workbook(template_file)
    today = datetime.today().strftime("%Y-%m-%d")

    df_full_purchase = pd.read_excel(purchase_file)
    df_mapping = pd.concat(pd.read_excel(mapping_file, sheet_name=None), ignore_index=True)
    df_mapping = df_mapping.dropna(subset=["1688原始商品名稱", "1688原始商品規格"])

    df_purchase = df_full_purchase[["货品标题", "数量", "运单号"]].rename(
        columns={"货品标题": "貨品標題", "数量": "數量", "运单号": "運單號"}
    )
    df_purchase["運單號"] = df_purchase["運單號"].apply(lambda x: str(int(x)) if pd.notna(x) else "")

    def fuzzy_lookup(name, spec, mapping_df):
        matches = difflib.get_close_matches(name, mapping_df["1688原始商品名稱"], n=1, cutoff=0.6)
        if matches:
            subset = mapping_df[mapping_df["1688原始商品名稱"] == matches[0]]
            spec_matches = difflib.get_close_matches(spec, subset["1688原始商品規格"], n=1, cutoff=0.6)
            if spec_matches:
                result = subset[subset["1688原始商品規格"] == spec_matches[0]]
                if not result.empty:
                    return result.iloc[0].to_dict()
        return {
            "1688原始商品名稱": name,
            "1688原始商品規格": spec,
            "蝦皮商品編號": pd.NA,
            "蝦皮商品名稱": "⚠️ 未對應",
            "蝦皮商品規格": pd.NA,
        }

    output_rows = []
    for _, row in df_purchase.iterrows():
        full_name = row["貨品標題"]
        if "颜色:" in full_name:
            name_part, spec_part = full_name.split("颜色:", 1)
            name, spec = name_part.strip(), spec_part.strip()
        else:
            name, spec = full_name.strip(), ""
        mapped = fuzzy_lookup(name, spec, df_mapping)
        mapped["數量"] = row["數量"]
        mapped["運單號"] = row["運單號"]
        output_rows.append(mapped)

    final_df = pd.DataFrame(output_rows).fillna("")

    if today in wb.sheetnames:
        del wb[today]
    ws = wb.create_sheet(title=today)

    for r in dataframe_to_rows(final_df, index=False, header=True):
        ws.append(r)

    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.protection.set_password("258085")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return {
        "success": True,
        "sheet": today,
        "preview": final_df,
        "report_bytes": output
    }
